from ScoreCalculator import ScoreCalculator
import openpyxl
import operator

# playter_score.xlsx파일 불러와서 데이터를 받아온다.
filename = "player_score.xlsx"

wb_obj = openpyxl.load_workbook(filename)
sheet_obj = wb_obj.active

max_row = 7    #행의 개수
max_col = 8     #열의 개수

currentCell_obj = sheet_obj.cell(row=1, column=1)
Score_list = []     #7명의 선수들의 점수 list
name_score = {}     #(선수, 평균점수) 쌍 딕셔너리

for i in range(2, max_row+1):       #player_score.xlsx 파일에서 '점수'와 '이름'을 가져온다.
    for j in range(1, max_col+1):
        currentCell_obj = sheet_obj.cell(row = i, column=j)
        Score_list.append(currentCell_obj.value)        # 셀의 value를 Score_list에 추가한다.

    #클래스 생성해서 값으로 받아온 list값들 넣어준다.
    Gymnast_cal = ScoreCalculator(Score_list[1], Score_list[2], Score_list[3], Score_list[4], Score_list[5], Score_list[6], Score_list[7])
    delMax_list = Gymnast_cal.del_max()  # 최대값 삭제
    delMin_list = Gymnast_cal.del_min()  # 최소값 삭제
    avg_score = Gymnast_cal.avg_score()  # 평균 구하기

    #이름-점수 쌍을 name_score에 삽입한다.
    name_score[Score_list[0]] = avg_score
    Score_list = []   # 한 행을 처리한 뒤 Score_list를 초기화 한다.

#받아온 name_score를 내림차순으로 정렬한다
# operator.itemgetter(1)로 하여 value값을 기준으로 한다.
#reverse = True로 하여 내림차순으로 출력되게 한다.
sortedScore = sorted(name_score.items(), key = operator.itemgetter(1), reverse=True)
sortedScoreList = list(sortedScore)        #list로 만들고

print("="*50)
print("Gymnast Awards Ceremony")
print("="*50)

#sortedScoreList에서 앞에 1,2,3 요소가 1,2,3 등이므로 출력(앞에서 정렬했기 때문에)
print("Gold medal : ", sortedScoreList[0][0], "(final score = ", sortedScoreList[0][1], ")")
print("Siver medal : ", sortedScoreList[1][0], "(final score = ", sortedScoreList[1][1], ")")
print("Bronze medal : ", sortedScoreList[2][0], "(final score = ", sortedScoreList[2][1], ")")

print("="*50)
print("Congratulations!")
